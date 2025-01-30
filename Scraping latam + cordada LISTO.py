import os
import re
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Directorios de los archivos de texto
txt_directory_cordada = r"C:\Users\Lucian\Documents\EMMA\txt cordada"
txt_directory_latam = r"C:\Users\Lucian\Documents\EMMA\txts"
# Ruta del archivo Excel de salida
excel_path = r"C:\Users\Lucian\Documents\EMMA\test excel\prueba.xlsx"

# Diccionarios para almacenar los datos y mapear las columnas
data_keys_cordada = {
    "Monto Operación": "Monto",
    "Fecha de Giro": "Fecha",
    "Anticipo": "Financiamiento",
    "Diferencia de Precio": "Dif Precio",
    "Gastos": "Gasto",
    "Tasa": "Tasa"
}

data_keys_latam = {
    "Fecha": "Fecha",
    "Monto total": "Monto",
    "Diferencia de precio": "Dif Precio",
    "IVA": "IVA",
    "Gastos y comisiones": "Gasto",
    "Tasa de mora": "Tasa"
}

# Función para extraer los datos de un archivo de texto
def extract_data(file_path, data_keys):
    with open(file_path, "r", encoding="latin-1") as file:
        content = file.read()
    
    extracted_data = {}
    for key in data_keys.keys():
        if key == "Fecha de Giro" or key == "Fecha":
            pattern = rf"{key}\s*:\s*(\d{{2}}/\d{{2}}/\d{{4}})"
        elif key == "Tasa" or key == "Tasa de mora":
            pattern = rf"{key}\s*:\s*([\d.,]+)\s*%?"
        elif key == "Anticipo":
            pattern = rf"{key}\s*:\s*([\d.,]+)\s*%"
        else:
            pattern = rf"{key}\s*:\s*(-?\$[\d.,]+)"
        match = re.search(pattern, content)
        extracted_data[key] = match.group(1) if match else None
    return extracted_data

# Cargar el archivo Excel existente
workbook = load_workbook(excel_path)
sheet = workbook.active

# Obtener los índices de las columnas basados en los encabezados
column_indices = {cell.value.strip(): col_idx for col_idx, cell in enumerate(sheet[1], 1)}

# Imprimir los encabezados encontrados en el archivo Excel
print("Encabezados encontrados en el archivo Excel:", list(column_indices.keys()))

# Verificar que todos los encabezados existen en el archivo Excel
for key in set(data_keys_cordada.values()).union(set(data_keys_latam.values())):
    if key not in column_indices:
        raise KeyError(f"El encabezado '{key}' no se encuentra en el archivo Excel.")

# Función para procesar archivos de texto y añadir filas al Excel
def process_files(txt_directory, data_keys, identifier):
    # Identificar la fila donde está el identificador en la columna "Fondo"
    fondo_col_idx = column_indices["Fondo"]
    identifier_row = None

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=fondo_col_idx, max_col=fondo_col_idx):
        for cell in row:
            if cell.value == identifier:
                identifier_row = cell.row
                break
        if identifier_row:
            break

    if not identifier_row:
        raise ValueError(f"No se encontró '{identifier}' en la columna 'Fondo'.")

    # Iterar sobre todos los archivos en el directorio de textos
    for filename in os.listdir(txt_directory):
        if filename.endswith(".txt"):
            txt_path = os.path.join(txt_directory, filename)
            extracted_data = extract_data(txt_path, data_keys)
            
            # Encontrar la primera fila vacía
            new_row_idx = sheet.max_row + 1
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                if all(cell.value is None for cell in row):
                    new_row_idx = row[0].row
                    break
            
            # Añadir la nueva fila con las fórmulas copiadas
            for col_idx in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=identifier_row, column=col_idx)
                target_cell = sheet.cell(row=new_row_idx, column=col_idx)
                if source_cell.has_style:
                    target_cell._style = source_cell._style
                if source_cell.data_type == 'f':
                    target_cell.value = source_cell.value.replace(str(identifier_row), str(new_row_idx))
                else:
                    target_cell.value = source_cell.value
            
            # Escribir los datos en las columnas correspondientes
            for key, value in extracted_data.items():
                col_idx = column_indices[data_keys[key]]
                if key == "Anticipo":
                    cell = sheet.cell(row=new_row_idx, column=col_idx)
                    cell.value = float(value.replace(',', '.')) / 100  # Convertir a porcentaje
                    cell.number_format = '0.0%'  # Formato de porcentaje
                else:
                    sheet.cell(row=new_row_idx, column=col_idx).value = value

# Procesar archivos de texto para "CORDADA"
process_files(txt_directory_cordada, data_keys_cordada, "CORDADA")
# Procesar archivos de texto para "LATAM"
process_files(txt_directory_latam, data_keys_latam, "LATAM")

# Asegurarse de que todos los valores en la columna "Financiamiento" tengan el formato correcto
financiamiento_col_idx = column_indices["Financiamiento"]
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=financiamiento_col_idx, max_col=financiamiento_col_idx):
    for cell in row:
        if cell.value is not None:
            cell.number_format = '0.0%'  # Formato de porcentaje

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print("Extracción y guardado en Excel completados.")