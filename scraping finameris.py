import os
import re
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Directorio de los archivos de texto
txt_directory = r"C:\Users\Lucian\Documents\EMMA\txt finameris"
# Ruta del archivo Excel de salida
excel_path = r"C:\Users\Lucian\Documents\EMMA\test excel\prueba.xlsx"

# Diccionario para almacenar los datos y mapear las columnas
data_keys = {
    "MONTO DOCUMENTOS": "Monto",
    "FECHA DE CURSE": "Fecha",
    "% DE ANTICIPO": "Financiamiento",
    "DIFERENCIA DE PRECIO": "Dif Precio",
    "GASTOS": "Gasto",
    "IMPUESTOS": "IVA"
}

# Función para extraer los datos de un archivo de texto
def extract_data(file_path):
    with open(file_path, "r", encoding="latin-1") as file:
        content = file.read()
    
    extracted_data = {}
    for key in data_keys.keys():
        if key == "FECHA DE CURSE":
            pattern = rf"{key}\s*:\s*(\d{{1,2}} de \w+ de \d{{4}})"
        elif key == "% DE ANTICIPO":
            pattern = rf"{key}\s*:\s*([\d.,]+)\s*%"
        else:
            pattern = rf"{key}\s*:\s*\$?\s*([\d.,]+)"
        match = re.search(pattern, content)
        extracted_data[key] = match.group(1) if match else None
    return extracted_data

# Cargar el archivo Excel existente
workbook = load_workbook(excel_path)
sheet = workbook.active

# Obtener los índices de las columnas basados en los encabezados
column_indices = {cell.value.strip(): col_idx for col_idx, cell in enumerate(sheet[1], 1)}

# Imprimir los encabezados encontrados en el archivo Excel
print("Encabezados encontrados en el archivo Excel:", list(column_indices.keys()))

# Verificar que todos los encabezados existen en el archivo Excel
for key in data_keys.values():
    if key not in column_indices:
        raise KeyError(f"El encabezado '{key}' no se encuentra en el archivo Excel.")

# Identificar la fila donde está "FINAMERIS" en la columna "Fondo"
fondo_col_idx = column_indices["Fondo"]
latam_row = None

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=fondo_col_idx, max_col=fondo_col_idx):
    for cell in row:
        if cell.value == "FINAMERIS":
            latam_row = cell.row
            break
    if latam_row:
        break

if not latam_row:
    raise ValueError("No se encontró 'FINAMERIS' en la columna 'Fondo'.")

# Iterar sobre todos los archivos en el directorio de textos
for filename in os.listdir(txt_directory):
    if filename.endswith(".txt"):
        txt_path = os.path.join(txt_directory, filename)
        extracted_data = extract_data(txt_path)
        
        # Añadir la nueva fila con las fórmulas copiadas
        new_row_idx = sheet.max_row + 1
        for col_idx in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=latam_row, column=col_idx)
            target_cell = sheet.cell(row=new_row_idx, column=col_idx)
            if source_cell.has_style:
                target_cell._style = source_cell._style
            if source_cell.data_type == 'f':
                target_cell.value = source_cell.value.replace(str(latam_row), str(new_row_idx))
            else:
                target_cell.value = source_cell.value
        
        # Escribir los datos en las columnas correspondientes
        for key, value in extracted_data.items():
            col_idx = column_indices[data_keys[key]]
            if value is not None:
                value = value.replace(',', '.')  # Reemplazar comas por puntos
            if key == "% DE ANTICIPO" and value is not None:
                cell = sheet.cell(row=new_row_idx, column=col_idx)
                cell.value = float(value) / 100  # Convertir a porcentaje
                cell.number_format = '0.0%'  # Formato de porcentaje
            else:
                sheet.cell(row=new_row_idx, column=col_idx).value = value

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print("Extracción y guardado en Excel completados.")